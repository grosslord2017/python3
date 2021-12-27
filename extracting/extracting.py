# задание
'''Ваша задача - автоматизировать процесс извлечения данных с itdashboard.gov.
Бот должен получить список агентств и сумму расходов с главной страницы.
Нажмите "DIVE IN" на главной странице, чтобы узнать суммы расходов для каждого агентства.
Запишите суммы в файл Excel и назовите таблицу «Агентства».
Затем бот должен выбрать одно из агентств, например, National Science Foundation
(это должно быть настроено в файле или в Robocloud).
Зайдя на страницу агентства, очистите таблицу со всеми «Индивидуальными инвестициями» и запишите ее на новый лист в Excel.
Если столбец «UII» содержит ссылку, откройте ее и загрузите PDF-файл с бизнес-обоснованием (кнопка «Загрузить
бизнес-кейс в формате PDF»)
Ваше решение должно быть отправлено и протестировано на Robocloud. -> свяжитесь с Андреем, чтобы получить доступ после
того, как код будет готов, или вы можете бесплатно создать свою учетную запись!
Сохраните загруженные файлы и лист Excel в корень выходной папки.'''

from selenium import webdriver
from selenium.webdriver.chrome.service import Service as ChromeService
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from time import sleep
import openpyxl
import os

# path_driver = '/home/chromedriver'
path_driver = os.path.dirname(os.path.realpath(__file__)) + '/chromedriver'
path = os.path.dirname(os.path.realpath(__file__)) + '/output/'
os.makedirs(path, exist_ok=True)
# name_agency = 'Department of Labor'
name_agency = 'Department of Education'
url_home = 'https://itdashboard.gov/'

# Chrome options
options = webdriver.ChromeOptions()
options.add_argument("--no-sandbox")
options.add_experimental_option("excludeSwitches", ["enable-automation"])
options.add_experimental_option("useAutomationExtension", False)
preferences = {"download.default_directory": path, "directory_upgrade": True, "safebrowsing.enabled": True }
options.add_experimental_option("prefs", preferences)
service = ChromeService(executable_path=path_driver)
driver = webdriver.Chrome(service=service, options=options)
driver.get(url_home)


# ---------
# выбор и создание списка агенств
def pull_off_values():
    agents = element.find_elements(By.CSS_SELECTOR, "a span.h4")
    agents_list = []
    for i in agents:
        agents_list.append(i.text)

    invests = element.find_elements(By.CSS_SELECTOR, "a span.h1")
    invests_list = []
    for i in invests:
        invests_list.append(i.text)

    btn = element.find_elements(By.CSS_SELECTOR, "a.btn")
    btn_list = []
    for i in btn:
        btn_list.append(i)

    # ------- если имя агенства которое нам надо, совпадает с именем в цикле перебора, то элемент кнопки этого
    # ------- агенства сохранить в переменную btn_elem
    result = []
    btn_elem = None
    for i in range(len(agents_list)):
        result.append([agents_list[i], invests_list[i]])
        if agents_list[i] == name_agency:
            btn_elem = btn_list[i]

    return result, btn_elem

# заполнение в Excel таблице страницы "агенства"
def add_agencys_table(agency_list):
    agency_l = agency_list
    excel_file = openpyxl.Workbook()
    excel_sheet = excel_file.create_sheet(title='агенства', index=0)
    del excel_file['Sheet']
    for i in range(len(agency_l)):
        excel_sheet[f'A{i + 1}'] = agency_l[i][0]
        excel_sheet[f'B{i + 1}'] = agency_l[i][1]
    excel_file.save(filename=f'{path}Агенства.xlsx')

def download_pdf(row):
    try:
        url = row.find_element(By.CSS_SELECTOR, "td.left.sorting_2 > a")
        link = url.get_attribute('href')
        s = f"window.open('{link}')"
        parent_handle = driver.window_handles[0]

        driver.execute_script(s)  # открывает новое окно в браузере

        child_handle = [x for x in driver.window_handles if x != parent_handle][0]
        driver.switch_to.window(child_handle)
        pdf_file = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, "div#business-case-pdf a")))
        onlyfiles_befor = next(os.walk(path))[2]
        count_files_befor = len(onlyfiles_befor)
        pdf_file.click()
        sleep(5)
        # блок проверки скачан файл или нет.
        while True:
            onlyfiles_after = next(os.walk(path))[2]
            count_files_after = len(onlyfiles_after)
            if count_files_befor == count_files_after:
                sleep(2)
                continue
            else:
                break

        # добавить проверку появился ли файл в директории, если да - следующая итерация, нет - ждем

        driver.close()
        driver.switch_to.window(parent_handle)
    except:
        None

# блок копирования таблицы в excel с выбором всей таблицы и одним проходом.
def add_table_of_agency():
    excel_file = openpyxl.load_workbook(path + 'Агенства.xlsx')

    if 'индивидуальные инвестиции' in excel_file.sheetnames:
        pfd = excel_file['индивидуальные инвестиции']
        excel_file.remove(pfd)
        excel_file.save("/home/uadmin/python3/testing/1.xlsx")

    excel_sheet_invest = excel_file.create_sheet(title='индивидуальные инвестиции', index=1)
    table = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CSS_SELECTOR, "div#investments-table-object_wrapper")))
    # ------- блок с выпадающим меню и выбором пункта ALL
    check_all_elem = table.find_elements(By.CSS_SELECTOR, "label select.form-control option")
    for el in check_all_elem:
        if el.text == 'All':
            el.click()
    # ------- ждем загрузку всей таблицы
    number_rows = table.find_element(By.CSS_SELECTOR, "div#investments-table-object_info").text
    n = number_rows.split()
    number_rows_str = f'{" ".join(n[0:3])} {n[5]} {" ".join(n[4:])}'
    WebDriverWait(table, 10).until(EC.text_to_be_present_in_element((By.CSS_SELECTOR, "div#investments-table-object_info"), number_rows_str))
    # --- ---
    rows = table.find_elements(By.CSS_SELECTOR, "tbody tr")
    r_iter = 1
    for row in rows:

        download_pdf(row)

        col_iter = 1
        for cells in row.find_elements(By.CSS_SELECTOR, "td"):
            excel_sheet_invest.cell(row=r_iter, column=col_iter, value=cells.text)
            col_iter += 1
        r_iter += 1
    excel_file.save(filename=f'{path}Агенства.xlsx')



# блок кликает на кнопку DIVE IN
cl = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "a[aria-controls='home-dive-in']")))
cl.click()
element = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CSS_SELECTOR, "div#agency-tiles-container")))
sleep(5)

agency, btn_elem = pull_off_values()
add_agencys_table(agency)
btn_elem.click()

# собирает данные на страницу 2
add_table_of_agency()

driver.close()

